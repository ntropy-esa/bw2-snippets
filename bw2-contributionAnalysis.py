#############################################
# bw2helpers.py contains a set of functions
# useful for creating LCA models in notebooks
# but also for contribution analysis and 
# plotting of results
#############################################

## Basic imports 
from __future__ import print_function, unicode_literals
import brightway2 as bw2
import numpy as np
import pandas as pd
import itertools
import json, datetime, pickle
import matplotlib.pyplot as plt

from lca_algebraic import newActivity, newFloatParam, getActByCode, persistParams
from bw2data.parameters import ActivityParameter, DatabaseParameter, ProjectParameter, Group
from eight import *
from bw2data import databases, methods, get_activity, Method
from bw2calc import LCA
from collections import defaultdict
from matplotlib import style
style.available
style.use('seaborn')


def ExportProjectParameters(f='ProjectParamExport.xlsx'):
    params = { i:p.dict for i,p in enumerate(ProjectParameter.select() )}
    pd.DataFrame.from_dict(params).T.to_excel('ProjectParamExport.xlsx')
    
def LoadProjectParameters(f='ProjectParamExport.xlsx'):
    df = pd.read_excel('ProjectParamExport.xlsx', index_col=0)
    paramDict = df.T.to_dict(orient='dict')
    paramList = [v for k,v in paramDict.items()]
    bw2.parameters.new_project_parameters(paramList, overwrite=True)
    bw2.parameters.recalculate()
    

    
## CONTRIBUTION ANALYSIS - Part 1: apply tags to databases
## Use v2 as it directly edits the databases, rather than re-writing it, does not cause peewee errors!
def conveniently_tag_database_v2(fg_db_to_tag = '', label='', ):
    '''
    Auxiliary function to conveniently assign new tag labels to a foreground database, for group analysis.
    Select a forground database to tag, via fg_db_to_tag;
    Then define the label name, via label
    Then, loop through all activities, and assign directly new attributes, using peewee functions to save

    Does not return anything. Changes are directly saved to database. Should avoid running into bugs of re-writing database. 
    
    Usage: conveniently_tag_database('fg_database', 'label_name')
    '''
    db = bw2.Database(fg_db_to_tag)
    print('There are %i items to be tagged, one by one' %(len(db)) )
    val = input("Do you want to proceed (Y/N) ? ")
    n=0
    r=0
    if val != 'Y':
        print('Okay, we stop here')
    else:
        print("Lets proceed! Type 'skip' or in order to not tag the given activity or remove existing tag")
        for act in db:
            if label in act:
                print('Activity: ', act['name'])
                print('Current group: ', act[label])
            else:
                print('Activity: ', act['name'])

            val = input("...to be in the group called... ? ")
            
            if val == 'skip':
                # need to pop the key if it was defined previously
                act.pop(label, 'label was not present')
                r+=1
            if val != 'skip':
                act[label] = val
                n+=1
            
            act.save() # save back to db
    print("Number of activities tagged: %i" %(n))
    print("Number of activities skipped: %i" %(r))
    
def rewrite_tagged_database(fg_db_to_tag, new_data):
    '''
    ESA: LEGACY, use not recommended, use instead: conveniently_tag_database_v2
    fg_db_to_dag : name of database to be re-written in bw2
    new_data : the newly tagged database
    bw2 will raise error if new data does not correspond to target database
    '''
    val = input("Do you want to proceed (Y/N) ? ")
    if val != 'Y':
        print('Okay, we stop here')
    else:
        db = bw2.Database(fg_db_to_tag)
        db.write(new_data)
        
        
def conveniently_tag_database(fg_db_to_tag = '', label='', ):
    '''
    ESA: LEGACY, use not recommended, use instead: conveniently_tag_database_v2
    Auxiliary function to conveniently assign new tag labels to a foreground database, for group analysis.
    Select a forground database to tag, via fg_db_to_tag;
    Then define the label name, via label
    Returns the tagged database as a new dictionnnary, to be checked, and then re-written in database
    
    Usage: new_data = conveniently_tag_database('fg_database', 'label_name')
    '''
    db = bw2.Database(fg_db_to_tag)
    data = db.load()
    new_data = {}
    print('There are %i items to be tagged, one by one' %(len(data)) )
    val = input("Do you want to proceed (Y/N) ? ")
    if val != 'Y':
        print('Okay, we stop here')
    else:
        print("Lets proceed! Type 'skip' in order to not tag the given activity")
        for pro_tpl, pro in data.items():
            val = input(pro['name'] + "... to be in the group called... ? ")
            
            if val == 'skip':
                # need to pop the key if it was defined previously
                pro.pop(label, 'label was not present')
            if val != 'skip':
                pro[label] = val
            
            new_data[pro_tpl] = pro      
    return new_data
    
## CONTRIBUTION ANALYSIS - Part 2: perform graph traversal & group results by taggs
## FOR 1 Impact category at a time
def traverse_tagged_databases(
    functional_unit, method, label="tag", default_tag="other", secondary_tags=[],
    fg_databases=None, bio2tech=False, parent4other=False
):

    """
    ESA : modify to allow performing grouping as in SimaPro, on multiple foreground databases, 
    with projection of biosphere emission on their technosphere activity, and propagation 
    of upstream tags through non-tag activities
    
    Traverse a functional unit throughout its foreground database(s) or the    
    listed databses in fg_databses, and group impacts by tag label.
    Contribution analysis work by linking impacts to individual activities.
    However, you also might want to group impacts in other ways. For example,
    give individual biosphere exchanges their own grouping, or aggregate two
    activities together.

    Consider this example system, where the letters are the tag labels, and the
    numbers are exchange amounts. The functional unit is one unit of the tree
    root.

    .. image:: images/tagged-traversal.png
       :alt: Example tagged supply chain

    In this supply chain, tags are applied to activities and biosphere exchanges.
    If a biosphere exchange is not tagged, it inherits the tag of its producing
    activity. Similarly, links to other databases are assessed with the usual
    LCA machinery, and the total LCA score is tagged according to its consuming
    activity. If an activity does not have a tag, a default tag is applied.

    We can change our visualization to show the use of the default tags:

    .. image:: images/tagged-traversal-2.png
       :alt: Example tagged supply chain

    And then we can manually calculate the tagged impacts. Normally we would
    need to know the actual biosphere flows and their respective
    characterization factors (CF), but in this example we assume that each
    CF is one. Our result, group by tags, would therefore be:

        * **A**: :math:`6 + 27 = 33`
        * **B**: :math:`30 + 44 = 74`
        * **C**: :math:`5 + 16 + 48 = 69`
        * **D**: :math:`14`

    This function will only traverse the foreground database, i.e. the
    database of the functional unit activity. A functional unit can have
    multiple starting nodes; in this case, all foreground databases are
    traversed.

    Input arguments:
        * ``functional_unit``: A functional unit dictionary, e.g. ``{("foo", "bar"): 42}``.
        * ``method``: A method name, e.g. ``("foo", "bar")``
        * ``label``: The label of the tag classifier. Default is ``"tag"``
        * ``default_tag``: The tag classifier to use if none was given. Default is ``"other"``
        * ``secondary_tags``: List of tuples in the format (secondary_label, secondary_default_tag). Default is empty list.
        
        * ``fg_databases``: a list of foreground databases to be traversed, e.g. ['foreground', 'biomass', 'machinery']
                            It's not recommended to include all databases of a project in the list to be traversed, especially not ecoinvent itself
        * ``bio2tech``: if set to ``True``, then any biosphere emissions arising in an activity x are given the same tag as the activity x.
        * ``parent4other``: if set to ``True``, untagged technosphere exchanges are aggregated with their parent .. become "outside"

    Returns:
       Aggregated tags dictionary from ``aggregate_tagged_graph``, and tagged supply chain graph from ``recurse_tagged_database``.
    """

    lca = LCA(functional_unit, method)
    lca.lci(factorize=True)
    lca.lcia()
    method_dict = {o[0]: o[1] for o in Method(method).load()}
    graph = [
        recurse_tagged_database(
            key, amount, method_dict, lca, label, default_tag, secondary_tags, fg_databases, parent4other
        )
        for key, amount in functional_unit.items()
    ]
    return aggregate_tagged_graph(graph, bio2tech), graph


def aggregate_tagged_graph(graph, bio2tech=False,):
    """Aggregate a graph produced by ``recurse_tagged_database`` by the provided tags.
    Outputs a dictionary with keys of tags and numeric values.  
    If bio2tech is set to True, then biosphere exchanges are added to the tag of the parent activity (instead of direct emissions)
    .. code-block:: python
        {'a tag': summed LCIA scores}

    """

    def recursor(obj, scores):
        scores[obj["tag"]] += obj["impact"]
        if bio2tech:
            for flow in obj["biosphere"]:
                scores[obj["tag"]] += flow["impact"]
        else: # default behavior
            for flow in obj["biosphere"]:
                scores[flow["tag"]] += flow["impact"]
        for exc in obj["technosphere"]:
            scores = recursor(exc, scores)
        return scores

    scores = defaultdict(int)
    for obj in graph:
        scores = recursor(obj, scores)
    return scores


def recurse_tagged_database(
    activity, amount, method_dict, lca, label, default_tag, secondary_tags=[], fg_databases=None, parent4other=False
):

    """Traverse a foreground database and assess activities and biosphere flows by tags.

    Input arguments:

        * ``activity``: Activity tuple or object
        * ``amount``: float
        * ``method_dict``: Dictionary of biosphere flow tuples to CFs, e.g. ``{("biosphere", "foo"): 3}``
        * ``lca``: An ``LCA`` object that is already initialized, i.e. has already calculated LCI and LCIA with same method as in ``method_dict``
        * ``label``: string
        * ``default_tag``: string
        * ``secondary_tags``: List of tuples in the format (secondary_label, secondary_default_tag). Default is empty list.      
        
        * ``fg_databases``: a list of foreground databases to be traversed, e.g. ['foreground', 'biomass', 'machinery']
                            It's not recommended to include all databases of a project in the list to be traversed, especially not ecoinvent itself
        * parent4other=False : if True, untagged technosphere exchanges are aggregated with their parent .. become "outside"

  Returns:

    .. code-block:: python

        {
            'activity': activity object,
            'amount': float,
            'tag': string,
            'secondary_tags': [list of strings],
            'impact': float (impact of inputs from outside foreground database),
            'biosphere': [{
                'amount': float,
                'impact': float,
                'tag': string,
                'secondary_tags': [list of strings]
            }],
            'technosphere': [this data structure]
        }
    """

    if isinstance(activity, tuple):
        activity = get_activity(activity)
        
    if fg_databases == None: # then set the list equal to the database of the functional unit 
    
        fg_databases = [activity['database']] # list, single item
    
    elif fg_databases == list(bw2.Database(activity['database']).find_graph_dependents()): 
        # check that the list fg_databases does not include all the databases involved in the FU 
        # (otherwise, it would mean we are likely to have to recurse through ecoinvent... not funny)
        # ideally, should only on first call of recurse_tagged_database
        raise Exception('The list of databases to traverse fg_databases should not be equal to the all databases involved in the project. You risk to attempt to traverse a background database like ecoinvent - it would take too much time')

    inputs = list(activity.technosphere())
    #print('activity', activity['name'])
    #print('inputs', inputs)
    
    production = list(activity.production())
    if len(production) == 1:
        scale = production[0]["amount"]
    elif not production:
        # Assume production amount of 1
        scale = 1
    else:
        raise ValueError("Can't scale by production exchange")

    inside = [exc for exc in inputs if exc["input"][0] in fg_databases] # inside = activities in fg_databases
    
    outside = {
        exc["input"]: exc["amount"] / scale * amount
        for exc in inputs
        if exc["input"][0] not in fg_databases ## calculates impacts for activities outside of fg_databases
    } # this is a dict of functional units, ready for lca score calculation

    if outside:
        lca.redo_lcia(outside)
        outside_score = lca.score
    else:
        outside_score = 0

    if parent4other:
        #if this option is set to True, will change default_tag's value to the tag
        # of the parent activity if itself was not empty 
        if activity.get(label) != None:
            default_tag = activity.get(label)
    
    if default_tag == 'Other':
        print(default_tag, amount, activity, outside_score)
        
    return {
        "activity": activity,
        "amount": amount,
        "tag": activity.get(label) or default_tag,
        "secondary_tags": [activity.get(t[0]) or t[1] for t in secondary_tags],
        "impact": outside_score,
        "biosphere": [
            {
                "amount": exc["amount"] / scale * amount,
                "impact": exc["amount"]
                / scale
                * amount
                * method_dict.get(exc["input"], 0),
                "tag": exc.get(label) or activity.get(label) or default_tag,
                "secondary_tags": [
                    exc.get(t[0]) or activity.get(t[0]) or t[1] for t in secondary_tags
                ],
            }
            for exc in activity.biosphere()
        ],
        "technosphere": [
            recurse_tagged_database(
                exc.input,
                exc["amount"] / scale * amount,
                method_dict,
                lca,
                label,
                default_tag,
                secondary_tags,
                fg_databases,
                parent4other
            )
            for exc in inside
        ],
    }

def recurse_for_search(activity, amount, searched, fg_databases=None):

    """
    For a given activity and amount, traverse a set of foreground databases
    looking for list of technosphere exchanges, building a graph of where & how 
    much of these activities are consumed.
    Is useful e.g. to check when & where is "transportation" consumed
    
    activity: key or bw2.activity object
    searched: list of keys or list of bw2.activity objects
    fg_databases: list of strings, names of databases
    """

    if isinstance(activity, tuple):
        activity = get_activity(activity) # activity as bw2 object
    if not isinstance(searched, list):
        searched = [searched] # if it was a single item, make a list of it
    if isinstance(searched[0], tuple):
        searched = [get_activity(s) for s in searched] # searched as list of bw2 object 
    
    if fg_databases == None: # then set the list equal to the database of the functional unit 
        fg_databases = [activity['database']] # list, single item
    
    elif fg_databases == list(bw2.Database(activity['database']).find_graph_dependents()): 
        # check that the list fg_databases does not include all the databases involved in the FU 
        # (otherwise, it would mean we are likely to have to recurse through ecoinvent... not funny)
        # ideally, should only on first call of recurse_tagged_database
        raise Exception('The list of databases to traverse fg_databases should not be equal to the all databases involved in the project. You risk to attempt to traverse a background database like ecoinvent - it would take too much time')
    
    inputs = list(activity.technosphere())
    production = list(activity.production())
    if len(production) == 1:
        scale = production[0]["amount"]
    elif not production:
        # Assume production amount of 1
        scale = 1
    else:
        raise ValueError("Can't scale by production exchange")
               
    # inside = activities to further recurse on
    inside = [exc for exc in inputs if exc['input'] not in searched and exc["input"][0] in fg_databases]

    # outside = activities in fg_databases & in the list of searched
    outside = {
        exc['input']: (exc["amount"], exc["amount"] / scale * amount) 
        for exc in inputs
        if exc['input'] in searched
    } # this is a dict of functional units, ready for lca score calculation
   
    return {
        "activity": activity,
        "amount": amount, # amount of that activity consumed
        "searched": {get_activity(exc):{'tkm_fu': val[1], 'tkm_raw': val[0]} 
                     for exc, val in outside.items()
                    },
        "downstream": [
            recurse_for_search(
                exc.input,
                exc["amount"] / scale * amount,
                searched,
                fg_databases,
            )
            for exc in inside
        ],
    }

def cumprint_searched_graph(graph, searched, methods=None, typ='tkm_fu', cumsum = {}, c=1, unit='tkm'):
    '''
    Both prints the graph and computes cumulative sum
    impact: tuple refering to an impact category
    '''
    
    def print_searched_graph(graph, typ='tkm_fu', cumsum = {}, c=1, unit='tkm'):
        ''' Text print of a graph generated by recurse_for_search 
            Recursive function
        '''
        s = '|'+'-'*c
        if c == 1:
            print(graph['activity']['name'])
        if isinstance(graph, dict): # first call, graph is a dict, with the fu
            if graph['searched']: # if not empty
                print(s, graph['activity']['name']) 
                for k, v in graph['searched'].items():
                    if np.abs(v[typ]) > 1e-15:
                        if unit==None:
                            print(s+'-+', "{0:0.3f}".format(v[typ]), k['unit'], k['name'])
                        else:
                            print(s+'-+', "{0:0.3f}".format(v[typ]), unit, k['name'])
                    if k['name'] in cumsum:
                        cumsum[k['name']] += v[typ]
                    else:
                        cumsum[k['name']] = v[typ]
                print('|') 
            a = graph['downstream'] # list of downstream activities

            for item in a:
                cumsum = print_searched_graph(item, typ, cumsum, c+1, unit)
        
        return cumsum
                
    cumsum = {}
    if c == 1:
        cumsum = print_searched_graph(graph, typ, cumsum, c, unit)
    
    # impact associated with the analysed flows
    def _multiLCA(activities, methods):
        """Simple wrapper around brightway API"""
        bw2.calculation_setups['process'] = {'inv': activities, 'ia': methods}
        lca = bw2.MultiLCA('process')
        cols = [act for act_amount in activities for act, amount in act_amount.items()]
        return pd.DataFrame(lca.results.T, index=[method for method in methods], columns=cols)
    
    if methods: # not None
        # fu for each searched activity as a list
        activities = [ {bg:1} for bg in searched] 
        # calculate EF for each activity, returns a df
        EFs = _multiLCA(activities, methods).T
        
            
    return pd.DataFrame.from_dict(
            {'amount':{bw2.get_activity(index)['name']+', '+bw2.get_activity(index)['unit']: cumsum[bw2.get_activity(index)['name']] for index, row in EFs.iterrows() if bw2.get_activity(index)['name'] in cumsum}, # cumsum is a dict {'activity name': value}
            'unitary impact': {bw2.get_activity(index)['name']+', '+bw2.get_activity(index)['unit']:row[methods[0]] for index, row in EFs.iterrows() if bw2.get_activity(index)['name'] in cumsum}, # 
            'total impact': {bw2.get_activity(index)['name']+', '+bw2.get_activity(index)['unit']:row[methods[0]]*cumsum[bw2.get_activity(index)['name']] for index, row in EFs.iterrows() if bw2.get_activity(index)['name'] in cumsum},
                                  }, orient='index').T

#### MULTI IMPACT CATEGORY RECURSE
## FOR multiple impact category at a time

def multi_traverse_tagged_databases(
    functional_unit, methods, label="tag", default_tag="other", secondary_tags=[],
    fg_databases=None, bio2tech=False, parent4other=False
):

    """
    ESA : modify to allow performing grouping as in SimaPro, on multiple foreground databases, 
    with projection of biosphere emission on their technosphere activity, and propagation 
    of upstream tags through non-tag activities
    
    Traverse a functional unit throughout its foreground database(s) or the    
    listed databses in fg_databses, and group impacts by tag label.
    Contribution analysis work by linking impacts to individual activities.
    However, you also might want to group impacts in other ways. For example,
    give individual biosphere exchanges their own grouping, or aggregate two
    activities together.

    Consider this example system, where the letters are the tag labels, and the
    numbers are exchange amounts. The functional unit is one unit of the tree
    root.

    .. image:: images/tagged-traversal.png
       :alt: Example tagged supply chain

    In this supply chain, tags are applied to activities and biosphere exchanges.
    If a biosphere exchange is not tagged, it inherits the tag of its producing
    activity. Similarly, links to other databases are assessed with the usual
    LCA machinery, and the total LCA score is tagged according to its consuming
    activity. If an activity does not have a tag, a default tag is applied.

    We can change our visualization to show the use of the default tags:

    .. image:: images/tagged-traversal-2.png
       :alt: Example tagged supply chain

    And then we can manually calculate the tagged impacts. Normally we would
    need to know the actual biosphere flows and their respective
    characterization factors (CF), but in this example we assume that each
    CF is one. Our result, group by tags, would therefore be:

        * **A**: :math:`6 + 27 = 33`
        * **B**: :math:`30 + 44 = 74`
        * **C**: :math:`5 + 16 + 48 = 69`
        * **D**: :math:`14`

    This function will only traverse the foreground database, i.e. the
    database of the functional unit activity. A functional unit can have
    multiple starting nodes; in this case, all foreground databases are
    traversed.

    Input arguments:
        * ``functional_unit``: A functional unit dictionary, e.g. ``{("foo", "bar"): 42}``.
        * ``methods``: A list of methods name, e.g. ``("foo", "bar")``
        * ``label``: The label of the tag classifier. Default is ``"tag"``
        * ``default_tag``: The tag classifier to use if none was given. Default is ``"other"``
        * ``secondary_tags``: List of tuples in the format (secondary_label, secondary_default_tag). Default is empty list.
        
        * ``fg_databases``: a list of foreground databases to be traversed, e.g. ['foreground', 'biomass', 'machinery']
                            It's not recommended to include all databases of a project in the list to be traversed, especially not ecoinvent itself
        * ``bio2tech``: if set to ``True``, then any biosphere emissions arising in an activity x are given the same tag as the activity x.
        * ``parent4other``: if set to ``True``, untagged technosphere exchanges are aggregated with their parent .. become "outside"

    Returns:
       Aggregated tags dictionary from ``aggregate_tagged_graph``, and tagged supply chain graph from ``recurse_tagged_database``.
    """

    lca = LCA(functional_unit, methods[0])
    lca.lci(factorize=True)
    lca.lcia()
    method_dicts = [{o[0]: o[1] for o in Method(method).load()} for method in methods]
    graph = [multi_recurse_tagged_database(key, amount, methods, method_dicts, lca, label, default_tag, secondary_tags, fg_databases, parent4other)
             for key, amount in functional_unit.items()]
    return multi_aggregate_tagged_graph(graph, bio2tech), graph


def multi_aggregate_tagged_graph(graph, bio2tech=False):
    """Aggregate a graph produced by ``recurse_tagged_database`` by the provided tags.
    Outputs a dictionary with keys of tags and numeric values.  
    If bio2tech is set to True, then biosphere exchanges are added to the tag of the parent activity (instead of direct emissions)
    .. code-block:: python
        {'a tag': summed LCIA scores}

    """

    def recursor(obj, scores):
        if not scores.get(obj['tag']):
            scores[obj['tag']] = [x for x in obj['impact']]
        else:
            scores[obj['tag']] = [sum(x) for x in zip(scores[obj['tag']], obj['impact'])]
        
        if bio2tech:
            for flow in obj["biosphere"]:
                if not scores.get(flow['tag']):
                    scores[obj["tag"]] = [x for x in flow["impact"] ] 
                else:
                    scores[obj["tag"]] = [sum(x) for x in zip(scores[flow['tag']], flow['impact'])]
                    
        else: # default behavior
            for flow in obj["biosphere"]:
                if not scores.get(flow['tag']):
                    scores[flow['tag']] = [x for x in flow['impact']]
                else:
                    scores[flow['tag']] = [sum(x) for x in zip(scores[flow['tag']], flow['impact'])]
        
        for exc in obj["technosphere"]:
            scores = recursor(exc, scores)
            
        return scores

    scores = defaultdict(int)
    for obj in graph:
        scores = recursor(obj, scores)
    return scores


def multi_recurse_tagged_database(
    activity, amount, methods, method_dicts, lca, label, default_tag, secondary_tags=[], fg_databases=None, parent4other=False
):

    """Traverse a foreground database and assess activities and biosphere flows by tags.

    Input arguments:

        * ``activity``: Activity tuple or object
        * ``amount``: float
        * ``method``: a list of methods
        * ``method_dicts``: a list of Dictionary of biosphere flow tuples to CFs, e.g. ``{("biosphere", "foo"): 3}``
        * ``lca``: An ``LCA`` object that is already initialized, i.e. has already calculated LCI and LCIA with same method as in ``method_dict``
        * ``label``: string
        * ``default_tag``: string
        * ``secondary_tags``: List of tuples in the format (secondary_label, secondary_default_tag). Default is empty list.      
        
        * ``fg_databases``: a list of foreground databases to be traversed, e.g. ['foreground', 'biomass', 'machinery']
                            It's not recommended to include all databases of a project in the list to be traversed, especially not ecoinvent itself
        * parent4other=False : if True, untagged technosphere exchanges are aggregated with their parent .. become "outside"

  Returns:

    .. code-block:: python

        {
            'activity': activity object,
            'amount': float,
            'tag': string,
            'secondary_tags': [list of strings],
            'impact': float (impact of inputs from outside foreground database),
            'biosphere': [{
                'amount': float,
                'impact': float,
                'tag': string,
                'secondary_tags': [list of strings]
            }],
            'technosphere': [this data structure]
        }
    """

    if isinstance(activity, tuple):
        activity = get_activity(activity)
        
    if fg_databases == None: # then set the list equal to the database of the functional unit 
        fg_databases = [activity['database']] # list, single item
    elif fg_databases == list(bw2.Database(activity['database']).find_graph_dependents()): 
        # check that the list fg_databases does not include all the databases involved in the FU 
        # (otherwise, it would mean we are likely to have to recurse through ecoinvent... not funny)
        # ideally, should only on first call of recurse_tagged_database
        raise Exception('The list of databases to traverse fg_databases should not be equal to the all databases involved in the project. You risk to attempt to traverse a background database like ecoinvent - it would take too much time')

    inputs = list(activity.technosphere())  
    production = list(activity.production())
    if len(production) == 1:
        scale = production[0]["amount"]
    elif not production:
        # Assume production amount of 1
        scale = 1
    else:
        raise ValueError("Can't scale by production exchange")

    inside = [exc for exc in inputs if exc["input"][0] in fg_databases] # inside = activities in fg_databases
    
#    outside = {
#        exc["input"]: exc["amount"] / scale * amount
#        for exc in inputs
#        if exc["input"][0] not in fg_databases ## calculates impacts for activities outside of fg_databases
#    } # this is a dict of functional units, ready for lca score calculation
    ## OBS: this creates a FU with all the BG activities and their amount, and the lca.score will be the sum of all of these
    ## BUT: if several inputs are the same, only one of them will be considered, since dict cannot repeat entries
    ## upgrade: check if already seen, and either add new or sum to existing amoutn
    outside = {}
    for exc in inputs:
        if exc["input"][0] not in fg_databases:
            if exc["input"] not in outside:
                outside[ exc["input"] ] = exc["amount"] / scale * amount
            else:
                outside[ exc["input"] ] += exc["amount"] / scale * amount
    
    
    if outside:
        outside_scores = []
        for n,m in enumerate(methods):
            #print("Switching to method ", n)
            lca.switch_method(m)
            lca.redo_lcia(outside)
            outside_scores.append(lca.score)
    else:
        outside_scores = [0]*len(methods)

    if parent4other:
        #if this option is set to True, will change default_tag's value to the tag
        # of the parent activity if itself was not empty 
        if activity.get(label) != None:
            default_tag = activity.get(label)
    
   # if default_tag == 'Other':
   #    print(default_tag, amount, activity, outside_scores)
        
    return {
        "activity": activity,
        "amount": amount,
        "tag": activity.get(label) or default_tag,
        "secondary_tags": [activity.get(t[0]) or t[1] for t in secondary_tags],
        "impact": outside_scores, ## ESA
        "biosphere": [
            {
                "amount": exc["amount"] / scale * amount,
                "impact": [ exc["amount"]
                / scale
                * amount
                * method_dict.get(exc["input"], 0) for method_dict in method_dicts],
                "tag": exc.get(label) or activity.get(label) or default_tag,
                "secondary_tags": [
                    exc.get(t[0]) or activity.get(t[0]) or t[1] for t in secondary_tags
                ],
            }
            for exc in activity.biosphere()
        ],
        "technosphere": [
            multi_recurse_tagged_database(
                exc.input,
                exc["amount"] / scale * amount,
                methods,
                method_dicts,
                lca,
                label,
                default_tag,
                secondary_tags,
                fg_databases,
                parent4other
            )
            for exc in inside
        ],
    }
### END MULTI RECURSE


def run_graphTaggedTraversal(fus, methods, methods_units, label, default_tag, fg_dbs, bio2tech=True, parent4other=True):
    '''
    For a set of functional units, and a set of impact assessment method
    # Check https://github.com/pjamesjoyce/lcopt/blob/development/lcopt/multi_tagged.py for a faster alternative using lca.switch_method()
    # 
    '''
    useFastMulti=True
  
    result_pds = pd.DataFrame()
    if not useFastMulti:
        for fu in fus:
            #lci, lcia
            a = bw2.get_activity( list(fu.items())[0][0] ) # tuple of activity
            print(a['name'], str(list(fu.items())[0][1]))
            scores = []
            for n, m in enumerate(methods): # inefficient on methods, should use multi-method version with switch lcia
                agg_graph, graph = traverse_tagged_databases(fu, m, 
                                                             label=label, default_tag=default_tag,
                                                             fg_databases=fg_dbs,
                                                             bio2tech=bio2tech,
                                                             parent4other=parent4other
                                                            )
                
                # dump the graph in some file
                now = datetime.datetime.now().strftime("%Y-%m-%d_at_%H-%M")
                gr_fp = 'graphs/'+label+'_'+a['name']+'_'+str(n)+'_'+now+'.pickle'
                print(gr_fp)
                gr_object = {'fu':fu, 'method':m, 'label':label, 'graph': graph}
                with open(gr_fp, 'wb') as outfile:
                    pickle.dump(gr_object, outfile)
                
                result_pd = pd.DataFrame(agg_graph, index=[0])

                result_pd['Impact'] = [m]
                result_pd['Units'] = [methods_units[n]]
                           
                result_pd['FU'] = [ a['name'] ]
                result_pd['FU_amount'] = [ str(list(fu.items())[0][1]) ]
                # list(fu.items())[0][0]+'-'+str(list(fu.items())[0][1]) 

                result_pds = result_pds.append(result_pd, sort=False)
                
    else: # fast version of the function, upgrade
        for fu in fus:
            a = bw2.get_activity( list(fu.items())[0][0] ) # tuple of activity
            print(a['name'], str(list(fu.items())[0][1]))
            scores = []
            agg_graph, graph = multi_traverse_tagged_databases(fu, methods, 
                                                             label=label, default_tag=default_tag,
                                                             fg_databases=fg_dbs,
                                                             bio2tech=bio2tech,
                                                             parent4other=parent4other
                                                             )
            # here agg_graph is a single dictionary shaped like {'group tag': [list of impact values for each method], 'next tag': [values]}
            result_pd = pd.DataFrame(agg_graph)
            result_pd['Impact'] = methods
            result_pd['Units'] = methods_units
            result_pd['FU'] = [ a['name'] ]*len(methods)
            result_pd['FU_amount'] = [ str(list(fu.items())[0][1]) ]*len(methods)
            
            result_pds = result_pds.append(result_pd, sort=False)
    
    result_pds.set_index(['FU','FU_amount', 'Impact', 'Units'], inplace=True)
    
    return result_pds
    
    
## PLOTTING & EXCEL EXPORTS
def barv(df, names, xlabel=None, ylabel=None, figsize=(10, 10), colormap='tab20b', width=0.7):
    '''
    USE:
    f, a = barv(sub_tmp, names, 
                'Biochar supply chain',
                r'kg $CO_2$-eq per kg biochar',
                (8,8),
                'Set2', # https://matplotlib.org/3.1.1/gallery/color/colormap_reference.html
                0.6)
    '''
    
    fig, axes = plt.subplots(nrows=1, ncols=1, figsize=figsize)

    df.plot(kind='bar', stacked=True,
         colormap=colormap, # viridis Pastel2 Paired tab20c https://matplotlib.org/3.1.1/gallery/color/colormap_reference.html 
         rot=0,
         width=width,
         ax=axes)

    #axes.set_yticklabels([act_name], fontsize=14)
    axes.set_xticklabels(names, fontsize=14)
    axes.set_xlabel(xlabel, fontsize=14)
    axes.tick_params(axis='x', which='major', labelsize=12)
    
    axes.set_ylabel(ylabel, fontsize=18)
    
    axes.legend(fontsize=14, loc='upper left', bbox_to_anchor=(1, 1.))
    axes.axhline(c='black')

    
    total_score = df.sum(axis=1)
    axes.plot(np.arange(0,df.shape[0],1), total_score ,'X', color='black', markersize=15)
    
    fig_handles = (fig, axes)
    print(names, total_score)
    return fig_handles

def barv_toAxes(ax, df, names, xlabel=None, ylabel=None, ticksize=14, colormap='tab20b', width=0.7, noCsink=True):
    '''
    Plot to the axe specified by ax
    '''
    df.plot(kind='bar', stacked=True,
     colormap=colormap, # viridis Pastel2 Paired tab20c https://matplotlib.org/3.1.1/gallery/color/colormap_reference.html 
     rot=0,
     width=width,
     ax=ax)

    #axes.set_yticklabels([act_name], fontsize=14)
    ax.set_xticklabels(names, fontsize=ticksize)
    ax.set_xlabel(xlabel, fontsize=14)
    ax.tick_params(axis='x', which='major', labelsize=ticksize)
    
    ax.set_ylabel(ylabel, fontsize=18)
    
    ax.legend(fontsize=14, loc='upper left', bbox_to_anchor=(1, 1.)).set_visible(False)
    ax.axhline(c='black')

    total_score = df.sum(axis=1)
    ax.plot(np.arange(0,df.shape[0],1), total_score ,'X', color='black', markersize=12)
    
    if noCsink:
        if 'C-sink' in df.columns:
            total_score_noSink = total_score -  df['C-sink']
            ax.plot(np.arange(0,df.shape[0],1), total_score_noSink ,'.', color='black', markersize=12)
            # https://matplotlib.org/stable/api/markers_api.html
        if 'Biochar C sequestration' in df.columns:
            total_score_noSink = total_score -  df['Biochar C sequestration']
            ax.plot(np.arange(0,df.shape[0],1), total_score_noSink ,'.', color='black', markersize=12)
            # https://matplotlib.org/stable/api/markers_api.html
            
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=True, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # re-set startrow
            startrow=0
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)
        
        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

## Grouped bar plot, with bars in legend
def add_line(ax, xpos, ypos, color='dimgrey'):
    line = plt.Line2D([xpos, xpos], [ypos + .1, ypos],
                      transform=ax.transAxes, linewidth=1, color=color,)
    line.set_clip_on(False)
    ax.add_line(line)

def label_len(my_index,level):
    labels = my_index.get_level_values(level)
    return [(k, sum(1 for i in g)) for k,g in itertools.groupby(labels)]

def label_group_bar_table(ax, df, font=[8, 12], color='dimgrey'):
    ypos = -.1
    scale = 1./df.index.size
    for level in range(df.index.nlevels)[::-1]:
        pos = 0
        for label, rpos in label_len(df.index,level):
            lxpos = (pos + .5 * rpos)*scale
            f = font[1] if rpos>1 else font[0]
            ax.text(lxpos, ypos, label, ha='center', transform=ax.transAxes, fontsize=f)
            add_line(ax, pos*scale, ypos, color)
            pos += rpos
        add_line(ax, pos*scale , ypos)
        ypos -= .1

## Grouped bar plot, with bars in legend
def add_line(ax, xpos, ypos, color='dimgrey'):
    line = plt.Line2D([xpos, xpos], [ypos + .1, ypos],
                      transform=ax.transAxes, linewidth=1, color=color,)
    line.set_clip_on(False)
    ax.add_line(line)

def add_line_h(ax, xpos, ypos, color='dimgrey'):
    line = plt.Line2D([xpos + .1, xpos], [ypos, ypos],
                      transform=ax.transAxes, linewidth=1, color=color,)
    line.set_clip_on(False)
    ax.add_line(line)
    
def label_len(my_index,level):
    labels = my_index.get_level_values(level)
    return [(k, sum(1 for i in g)) for k,g in itertools.groupby(labels)]

def label_group_bar_table2(ax, df, font=[8, 12], color='dimgrey'):
    ypos = -.1
    scale = 1./df.index.size
    for level in range(df.index.nlevels)[::-1]:
        pos = 0
        for label, rpos in label_len(df.index,level):
            lxpos = (pos + .5 * rpos)*scale
            
            f = font[1] if rpos>1 else font[0]

            ax.text(lxpos, ypos, label, ha='center', transform=ax.transAxes, fontsize=f)
            add_line(ax, pos*scale, ypos, color)
            
            if rpos>1:
                # impact label, add a line vertical
                xpos=pos*scale
                line = plt.Line2D([xpos, xpos], [0, 1],
                                  transform=ax.transAxes, linewidth=1, color=color)
                line.set_clip_on(False)
                ax.add_line(line)
                
            pos += rpos

        add_line(ax, pos*scale , ypos, color)
        ypos -= .1
    
    # 0-line
    ax.axhline(color='grey', linewidth=0.8)


def label_group_bar_table3(ax, df, font=[8, 12], color='dimgrey'):
    ''' For horizontal bars, labels on y-axis '''
    ypos = -.1
    scale = 1./df.index.size
    for level in range(df.index.nlevels)[::-1]:
        pos = 0
        for label, rpos in label_len(df.index,level):
            lxpos = (pos + .5 * rpos)*scale
            
            f = font[1] if rpos>1 else font[0]
            offset_x = 0.08 if rpos<=1 else 0.00
            offset_y = .0055 if rpos<=1 else -0.02
            rot = 0 if rpos<=1 else 90
            di = 'right' if rpos<=1 else 'center'
            ax.text(ypos+offset_x, lxpos-0.01+offset_y , label, ha=di, transform=ax.transAxes, fontsize=f, rotation=rot)
            add_line_h(ax, ypos ,pos*scale , color)
            
            if rpos>1:
                # impact label, add a line vertical
                xpos=pos*scale
                line = plt.Line2D([0, 1], [xpos, xpos],
                                  transform=ax.transAxes, linewidth=1, color=color)
                line.set_clip_on(False)
                ax.add_line(line)
                
            pos += rpos

        add_line_h(ax, ypos, pos*scale, color)
        ypos -= .1
    
    # 0-line
    ax.axvline(color='grey', linewidth=0.8)

    
def normalize_by_group(df, level=0):
    '''For impact categories
    level : int, which index level should we group by
    '''
    norm_map = {}
    for im in df.index.levels[level]: # is sorted... alphabetically ...
        ix = df.index.levels[level].name
        qr = str(ix)+' == "'+ str(im) + '"'
        m = np.max(np.abs(df.query(qr).sum(axis=1)))
        norm_map[im] = m
    
    for i, row in df.iterrows():
        df.loc[i,:] = df.loc[i,:] / norm_map[i[level]] * 100
    return df